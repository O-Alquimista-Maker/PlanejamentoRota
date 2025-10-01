# No início do arquivo, certifique-se de que 'flash' está importado
from flask import Flask, render_template, request, jsonify, redirect, url_for, send_file, flash
import calendar
from datetime import datetime
import json
import database_manager
import locale
import io
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

import sys
import os
import shutil
from waitress import serve

# --- LÓGICA DE CAMINHOS E CONFIGURAÇÃO INICIAL ---
def get_base_path():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    else:
        return os.path.abspath(".")

BASE_PATH = get_base_path()
DB_NAME = 'database.db'
DB_PATH = os.path.join(BASE_PATH, DB_NAME)
LOGO_NAME = 'logo.png'
LOGO_PATH = os.path.join(BASE_PATH, LOGO_NAME)

IS_FROZEN = getattr(sys, 'frozen', False)

if IS_FROZEN:
    template_folder = os.path.join(sys._MEIPASS, 'templates')
    static_folder = os.path.join(sys._MEIPASS, 'static')
    db_template_path = os.path.join(sys._MEIPASS, DB_NAME)
    if not os.path.exists(DB_PATH):
        shutil.copyfile(db_template_path, DB_PATH)
    app = Flask(__name__, template_folder=template_folder, static_folder=static_folder)
else:
    app = Flask(__name__)

app.secret_key = 'uma-chave-secreta-muito-segura'

database_manager.NOME_BANCO_DADOS = DB_PATH

try:
    locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')
except locale.Error:
    try:
        locale.setlocale(locale.LC_ALL, 'Portuguese_Brazil.1252')
    except locale.Error:
        print("Locale pt_BR não encontrado.")

# --- ROTAS DO CALENDÁRIO E API ---
@app.route('/')
def index_redirect():
    hoje = datetime.today()
    return redirect(url_for('calendario_view', ano=hoje.year, mes=hoje.month))

@app.route('/<int:ano>/<int:mes>')
def calendario_view(ano, mes):
    if mes < 1:
        return redirect(url_for('calendario_view', ano=ano - 1, mes=12))
    if mes > 12:
        return redirect(url_for('calendario_view', ano=ano + 1, mes=1))

    calendar.setfirstweekday(calendar.SUNDAY)

    # MODIFICAÇÃO: Busca os clientes e já ordena por nome (A-Z)
    clientes = sorted(database_manager.buscar_todos_clientes(), key=lambda c: c['nome_cliente'].lower())
    
    planejamento_mes = database_manager.buscar_planejamento_por_mes(ano, mes)
    planejamento_dict = {}
    for item in planejamento_mes:
        dia = item['dia']
        if dia not in planejamento_dict:
            planejamento_dict[dia] = []
        planejamento_dict[dia].append(item)
    
    cal = calendar.monthcalendar(ano, mes)
    
    mes_anterior, ano_anterior = (mes - 1, ano) if mes > 1 else (12, ano - 1)
    mes_proximo, ano_proximo = (mes + 1, ano) if mes < 12 else (1, ano + 1)
    
    return render_template('index.html', 
                           ano=ano, 
                           mes=mes, 
                           nome_mes=calendar.month_name[mes].capitalize(), 
                           calendario=cal, 
                           clientes=clientes, 
                           planejamento_json=json.dumps(planejamento_dict), 
                           nav_anterior={'ano': ano_anterior, 'mes': mes_anterior}, 
                           nav_proximo={'ano': ano_proximo, 'mes': mes_proximo}, 
                           gerenciar_clientes_url=url_for('gerenciar_clientes_view'))

@app.route('/api/copiar_mes_anterior/<int:ano>/<int:mes>')
def copiar_mes_anterior(ano, mes):
    mes_anterior, ano_anterior = (mes - 1, ano) if mes > 1 else (12, ano - 1)
    planejamento_anterior = database_manager.buscar_planejamento_por_mes(ano_anterior, mes_anterior)
    return jsonify(planejamento_anterior)

@app.route('/api/salvar_planejamento', methods=['POST'])
def salvar_planejamento():
    dados = request.json
    ano = dados.get('ano')
    mes = dados.get('mes')
    planejamento_do_mes = dados.get('planejamento')
    if not all([ano, mes, planejamento_do_mes is not None]):
        return jsonify({"status": "erro", "mensagem": "Dados incompletos recebidos."}), 400
    try:
        num_salvos = database_manager.salvar_planejamento_mes(ano, mes, planejamento_do_mes)
        nome_mes_pt = calendar.month_name[mes].capitalize()
        mensagem = f"{num_salvos} agendamentos salvos para {nome_mes_pt}/{ano}."
        return jsonify({"status": "sucesso", "mensagem": mensagem})
    except Exception as e:
        return jsonify({"status": "erro", "mensagem": f"Erro interno do servidor: {e}"}), 500

# --- ROTA DE EXPORTAÇÃO ---
@app.route('/exportar_excel/<int:ano>/<int:mes>')
def exportar_excel(ano, mes):
    try:
        calendar.setfirstweekday(calendar.SUNDAY)
        
        wb = Workbook()
        ws = wb.active
        nome_mes_pt = calendar.month_name[mes].capitalize()
        ws.title = f"{nome_mes_pt} {ano}"
        bold_font = Font(bold=True, size=10)
        normal_font = Font(size=10)
        center_align = Alignment(horizontal='center', vertical='center')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 10
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 25
        ws.row_dimensions[3].height = 25
        ws.row_dimensions[4].height = 20
        ws.merge_cells('A1:A3')
        if os.path.exists(LOGO_PATH):
            img = Image(LOGO_PATH)
            img.height = 70
            img.width = 140
            ws.add_image(img, 'A1')
        ws.merge_cells('B1:D3')
        ws['B1'] = f"Cronograma de Coleta - {nome_mes_pt}/{ano}"
        ws['B1'].alignment = center_align
        ws['B1'].font = Font(bold=True, size=12)
        ws['E1'] = "Código"
        ws['F1'] = "Data"
        ws['G1'] = "Revisão"
        ws.merge_cells('E2:E3')
        ws['E2'] = "F.082"
        ws.merge_cells('F2:F3')
        ws['F2'] = "22/03/2022"
        ws.merge_cells('G2:G3')
        ws['G2'] = "01"
        ws['A4'] = "Registros"
        ws['B4'] = "Atualizado em"
        ws['C4'] = datetime.now().strftime('%d/%m/%Y')
        ws['D4'] = "Responsável"
        ws['E4'] = "RICARDO"
        ws['F4'] = "Aprovado por"
        ws['G4'] = "RICARDO"
        for row in ws['A1':'G4']:
            for cell in row:
                cell.border = thin_border
                cell.alignment = center_align
                cell.font = normal_font
                if (cell.row == 1 and cell.column > 4) or cell.row == 4:
                    cell.font = bold_font
        ws['B1'].font = Font(bold=True, size=12)
        planejamento_mes = database_manager.buscar_planejamento_por_mes(ano, mes)
        clientes_db = {c['id']: c['nome_cliente'] for c in database_manager.buscar_todos_clientes()}
        dados_calendario = {}
        legenda_necessaria = False
        for item in planejamento_mes:
            dia = item['dia']
            equipe = item['equipe']
            cliente_id = item['id_cliente']
            cliente_nome = clientes_db.get(cliente_id, 'Cliente não encontrado')
            lab_externo = item.get('lab_externo', 0)
            if lab_externo:
                cliente_nome = f"★ {cliente_nome}"
                legenda_necessaria = True
            if (dia, equipe) not in dados_calendario:
                dados_calendario[(dia, equipe)] = []
            dados_calendario[(dia, equipe)].append(cliente_nome)
        cal = calendar.monthcalendar(ano, mes)
        dias_semana = ["Domingo", "Segunda", "Terça", "Quarta", "Quinta", "Sexta", "Sábado"]
        day_header_font = Font(bold=True, size=12)
        day_header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        cell_align_cal = Alignment(horizontal='left', vertical='top', wrap_text=True)
        r01_fill = PatternFill(start_color="E0F7FA", end_color="E0F7FA", fill_type="solid")
        r02_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
        ws.row_dimensions[5].height = 30
        for col, dia_nome in enumerate(dias_semana, 1):
            cell = ws.cell(row=5, column=col, value=dia_nome)
            cell.font = day_header_font
            cell.fill = day_header_fill
            cell.alignment = center_align
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col)].width = 25
        row_offset = 6
        for semana in cal:
            max_clientes_r1 = 1
            max_clientes_r2 = 1
            for dia in semana:
                if dia != 0:
                    max_clientes_r1 = max(max_clientes_r1, len(dados_calendario.get((dia, 'R1'), [])))
                    max_clientes_r2 = max(max_clientes_r2, len(dados_calendario.get((dia, 'R2'), [])))
            ws.row_dimensions[row_offset].height = 20
            ws.row_dimensions[row_offset + 1].height = (max_clientes_r1 + 1) * 15
            ws.row_dimensions[row_offset + 2].height = (max_clientes_r2 + 1) * 15
            for col, dia in enumerate(semana, 1):
                dia_cell = ws.cell(row=row_offset, column=col)
                r01_cell = ws.cell(row=row_offset + 1, column=col)
                r02_cell = ws.cell(row=row_offset + 2, column=col)
                dia_cell.border = thin_border
                r01_cell.border = thin_border
                r02_cell.border = thin_border
                if dia == 0:
                    ws.merge_cells(start_row=row_offset, start_column=col, end_row=row_offset + 2, end_column=col)
                    dia_cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                else:
                    dia_cell.value = dia
                    dia_cell.font = Font(bold=True)
                    clientes_r1 = dados_calendario.get((dia, 'R1'), [])
                    r01_cell.value = "R01:\n" + "\n".join(f"- {c}" for c in clientes_r1)
                    r01_cell.alignment = cell_align_cal
                    clientes_r2 = dados_calendario.get((dia, 'R2'), [])
                    r02_cell.value = "R02:\n" + "\n".join(f"- {c}" for c in clientes_r2)
                    r02_cell.alignment = cell_align_cal
                    if col != 1 and col != 7:
                        r01_cell.fill = r01_fill
                        r02_cell.fill = r02_fill
            row_offset += 3
        if legenda_necessaria:
            linha_legenda = row_offset
            ws.merge_cells(start_row=linha_legenda, start_column=1, end_row=linha_legenda, end_column=7)
            legenda_cell = ws.cell(row=linha_legenda, column=1)
            legenda_cell.value = "★ = Acompanhamento Laboratório Externo"
            legenda_cell.font = Font(bold=True, color="d32f2f")
            row_offset += 1
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        nome_arquivo = f"Planejamento_{nome_mes_pt}_{ano}.xlsx"
        return send_file(buffer, as_attachment=True, download_name=nome_arquivo, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        print(f"Erro detalhado ao gerar Excel: {e}")
        if isinstance(e, FileNotFoundError):
            return jsonify({"status": "erro", "mensagem": f"Erro ao gerar Excel: Arquivo '{LOGO_NAME}' não encontrado na pasta do projeto."}), 500
        return jsonify({"status": "erro", "mensagem": f"Erro ao gerar Excel: {str(e)}"}), 500

# --- ROTAS PARA GERENCIAMENTO DE CLIENTES ---

@app.route('/clientes')
def gerenciar_clientes_view():
    # MODIFICAÇÃO: Busca os clientes e já ordena por nome (A-Z)
    clientes = sorted(database_manager.buscar_todos_clientes(), key=lambda c: c['nome_cliente'].lower())
    return render_template('gerenciar_clientes.html', clientes=clientes)

@app.route('/clientes/adicionar', methods=['POST'])
def adicionar_cliente_action():
    nome = request.form.get('nome_cliente')
    endereco = request.form.get('endereco')
    telefone = request.form.get('telefone')
    if nome:
        database_manager.adicionar_cliente(nome, endereco, telefone)
        flash(f"Cliente '{nome}' adicionado com sucesso!", 'success')
    return redirect(url_for('gerenciar_clientes_view'))

@app.route('/clientes/editar/<int:id_cliente>', methods=['POST'])
def editar_cliente_action(id_cliente):
    nome = request.form.get('nome_cliente')
    endereco = request.form.get('endereco')
    telefone = request.form.get('telefone')
    if nome:
        database_manager.editar_cliente(id_cliente, nome, endereco, telefone)
        flash(f"Cliente '{nome}' atualizado com sucesso!", 'success')
    return redirect(url_for('gerenciar_clientes_view'))

@app.route('/clientes/excluir/<int:id_cliente>', methods=['POST'])
def excluir_cliente_action(id_cliente):
    database_manager.excluir_cliente(id_cliente)
    flash("Cliente excluído com sucesso!", 'success')
    return redirect(url_for('gerenciar_clientes_view'))

# --- PONTO DE ENTRADA DA APLICAÇÃO ---
if __name__ == '__main__':
    database_manager.inicializar() 
    if IS_FROZEN:
        print("Iniciando servidor de produção Waitress na porta 5000...")
        serve(app, host='127.0.0.1', port=5000)
    else:
        print("Iniciando servidor de desenvolvimento Flask...")
        # Linha correta
        app.run(debug=True, host='127.0.0.1', port=5000)

